from __future__ import annotations

from collections import defaultdict
from io import BytesIO
from typing import Literal

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.entities import Classroom, LessonSlot, ScheduleItem, School, StudentClass, Subject, Teacher

ExportView = Literal["class", "teacher", "school"]
ExportFormat = Literal["xlsx", "pdf"]

DAY_LABELS = {1: "Mon", 2: "Tue", 3: "Wed", 4: "Thu", 5: "Fri"}
DAY_ORDER = (1, 2, 3, 4, 5)

TITLE_FONT = Font(bold=True, size=14, color="1E3A5F")
SECTION_FONT = Font(bold=True, size=12, color="1E3A5F")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
META_FONT = Font(italic=True, color="666666", size=9)
CELL_FONT = Font(size=9)
CELL_ALIGNMENT = Alignment(wrap_text=True, vertical="top", horizontal="center")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center")
THIN_BORDER = Border(
    left=Side(style="thin", color="CCCCCC"),
    right=Side(style="thin", color="CCCCCC"),
    top=Side(style="thin", color="CCCCCC"),
    bottom=Side(style="thin", color="CCCCCC"),
)
ALT_ROW_FILL = PatternFill("solid", fgColor="F4F7FB")


def _collect_rows(db: Session, school_id: int, view: ExportView, entity_id: int | None) -> list[dict[str, object]]:
    items = list(db.scalars(select(ScheduleItem).where(ScheduleItem.school_id == school_id)))
    classes = {row.id: row.class_name for row in db.scalars(select(StudentClass).where(StudentClass.school_id == school_id))}
    teachers = {row.id: row.full_name for row in db.scalars(select(Teacher).where(Teacher.school_id == school_id))}
    rooms = {row.id: row.room_number for row in db.scalars(select(Classroom).where(Classroom.school_id == school_id))}
    subjects = {row.id: row.name for row in db.scalars(select(Subject))}
    slots = {row.id: row for row in db.scalars(select(LessonSlot))}

    if view == "class":
        if entity_id is None:
            raise ValueError("entity_id is required for class export")
        items = [item for item in items if item.class_id == entity_id]
    elif view == "teacher":
        if entity_id is None:
            raise ValueError("entity_id is required for teacher export")
        items = [item for item in items if item.teacher_id == entity_id]

    rows: list[dict[str, object]] = []
    for item in items:
        slot = slots.get(item.lesson_slot_id)
        rows.append(
            {
                "class_name": classes.get(item.class_id, f"#{item.class_id}"),
                "subject_name": subjects.get(item.subject_id, f"#{item.subject_id}"),
                "teacher_name": teachers.get(item.teacher_id, f"#{item.teacher_id}"),
                "room_number": rooms.get(item.classroom_id, f"#{item.classroom_id}"),
                "day_of_week": slot.day_of_week if slot else None,
                "lesson_number": slot.lesson_number if slot else None,
                "start_time": slot.start_time.strftime("%H:%M") if slot and slot.start_time else "",
                "end_time": slot.end_time.strftime("%H:%M") if slot and slot.end_time else "",
            }
        )
    rows.sort(
        key=lambda row: (
            str(row["class_name"]),
            row["day_of_week"] or 0,
            row["lesson_number"] or 0,
        )
    )
    return rows


def _group_rows_by_class(rows: list[dict[str, object]]) -> dict[str, list[dict[str, object]]]:
    grouped: dict[str, list[dict[str, object]]] = defaultdict(list)
    for row in rows:
        grouped[str(row["class_name"])].append(row)
    return dict(sorted(grouped.items(), key=lambda item: item[0]))


def _lesson_numbers_in_rows(rows: list[dict[str, object]]) -> list[int]:
    nums = {int(row["lesson_number"]) for row in rows if row.get("lesson_number") is not None}
    if not nums:
        return list(range(1, 8))
    return list(range(1, max(nums) + 1))


def _cell_text(row: dict[str, object]) -> str:
    time_part = ""
    if row.get("start_time") and row.get("end_time"):
        time_part = f"\n{row['start_time']}–{row['end_time']}"
    return f"{row['subject_name']}\n{row['teacher_name']}\nRoom {row['room_number']}{time_part}"


def _build_lesson_grid(rows: list[dict[str, object]]) -> dict[int, dict[int, str]]:
    """lesson_number -> day_of_week -> multiline cell text."""
    grid: dict[int, dict[int, str]] = defaultdict(dict)
    for row in rows:
        day = row.get("day_of_week")
        lesson = row.get("lesson_number")
        if day is None or lesson is None:
            continue
        grid[int(lesson)][int(day)] = _cell_text(row)
    return grid


def _style_range(ws, min_row: int, max_row: int, min_col: int, max_col: int, *, header_row: int | None = None) -> None:
    for row_idx in range(min_row, max_row + 1):
        for col_idx in range(min_col, max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = THIN_BORDER
            if header_row is not None and row_idx == header_row:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = HEADER_ALIGNMENT
            elif row_idx > (header_row or min_row) and col_idx == min_col:
                cell.font = Font(bold=True, size=9, color="1E3A5F")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.font = CELL_FONT
                cell.alignment = CELL_ALIGNMENT
                if row_idx > (header_row or min_row) and (row_idx - (header_row or min_row)) % 2 == 0:
                    cell.fill = ALT_ROW_FILL


def _write_class_timetable_block(
    ws,
    *,
    start_row: int,
    class_name: str,
    rows: list[dict[str, object]],
    subtitle: str | None = None,
) -> int:
    """Write a titled weekly grid; returns the next free row index."""
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=6)
    title_cell = ws.cell(row=start_row, column=1, value=f"Class {class_name}")
    title_cell.font = SECTION_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    row = start_row + 1
    if subtitle:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        meta = ws.cell(row=row, column=1, value=subtitle)
        meta.font = META_FONT
        row += 1

    header_row = row
    ws.cell(row=row, column=1, value="Lesson")
    for col_offset, day in enumerate(DAY_ORDER, start=2):
        ws.cell(row=row, column=col_offset, value=DAY_LABELS[day])
    row += 1

    grid = _build_lesson_grid(rows)
    for lesson in _lesson_numbers_in_rows(rows):
        ws.cell(row=row, column=1, value=lesson)
        for col_offset, day in enumerate(DAY_ORDER, start=2):
            ws.cell(row=row, column=col_offset, value=grid.get(lesson, {}).get(day, ""))
        row += 1

    _style_range(ws, header_row, row - 1, 1, 6, header_row=header_row)
    ws.row_dimensions[header_row].height = 22
    for lesson_row in range(header_row + 1, row):
        ws.row_dimensions[lesson_row].height = 52

    for col_idx in range(1, 7):
        letter = get_column_letter(col_idx)
        ws.column_dimensions[letter].width = 10 if col_idx == 1 else 22

    return row + 2


def _build_xlsx(
    grouped: dict[str, list[dict[str, object]]],
    *,
    document_title: str,
    layout: Literal["multi_sheet", "sections"],
    section_subtitle: str | None = None,
) -> bytes:
    wb = Workbook()
    if layout == "multi_sheet":
        default = wb.active
        wb.remove(default)
        if not grouped:
            ws = wb.create_sheet("Schedule")
            ws["A1"] = document_title
            ws["A1"].font = TITLE_FONT
            ws["A3"] = "No schedule items found."
        else:
            index_ws = wb.create_sheet("Classes", 0)
            index_ws["A1"] = document_title
            index_ws["A1"].font = TITLE_FONT
            index_ws["A3"] = "Class"
            index_ws["B3"] = "Lessons"
            index_ws["A3"].font = HEADER_FONT
            index_ws["B3"].font = HEADER_FONT
            index_ws["A3"].fill = HEADER_FILL
            index_ws["B3"].fill = HEADER_FILL
            idx_row = 4
            for class_name, class_rows in grouped.items():
                safe_title = class_name[:31]
                ws = wb.create_sheet(safe_title)
                _write_class_timetable_block(
                    ws,
                    start_row=1,
                    class_name=class_name,
                    rows=class_rows,
                    subtitle=section_subtitle,
                )
                index_ws.cell(row=idx_row, column=1, value=class_name)
                index_ws.cell(row=idx_row, column=2, value=len(class_rows))
                idx_row += 1
            index_ws.column_dimensions["A"].width = 18
            index_ws.column_dimensions["B"].width = 12
    else:
        ws = wb.active
        ws.title = "Schedule"
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        ws["A1"] = document_title
        ws["A1"].font = TITLE_FONT
        row = 3
        if not grouped:
            ws.cell(row=row, column=1, value="No schedule items found.")
        else:
            for class_name, class_rows in grouped.items():
                subtitle = section_subtitle
                row = _write_class_timetable_block(
                    ws,
                    start_row=row,
                    class_name=class_name,
                    rows=class_rows,
                    subtitle=subtitle,
                )

    payload = BytesIO()
    wb.save(payload)
    return payload.getvalue()


def _escape_pdf_text(value: object) -> str:
    return str(value).replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _grid_to_pdf_lines(class_name: str, rows: list[dict[str, object]], subtitle: str | None = None) -> list[str]:
    lines = [f"Class {class_name}"]
    if subtitle:
        lines.append(subtitle)
    lines.append("")
    header = "        " + "  ".join(f"{DAY_LABELS[d]:<18}" for d in DAY_ORDER)
    lines.append(header)
    lines.append("")

    grid = _build_lesson_grid(rows)
    for lesson in _lesson_numbers_in_rows(rows):
        parts = [f"L{lesson:<2}"]
        for day in DAY_ORDER:
            raw = grid.get(lesson, {}).get(day, "")
            first_line = raw.split("\n")[0] if raw else "—"
            parts.append(f"{first_line[:18]:<18}")
        lines.append("  ".join(parts))
        detail_lines: list[str] = []
        for day in DAY_ORDER:
            raw = grid.get(lesson, {}).get(day, "")
            if not raw:
                detail_lines.append("")
                continue
            chunks = raw.split("\n")
            teacher = chunks[1] if len(chunks) > 1 else ""
            detail_lines.append(teacher[:18])
        if any(detail_lines):
            lines.append("      " + "  ".join(f"{t:<18}" for t in detail_lines))
    lines.append("")
    return lines


def _build_pdf(
    grouped: dict[str, list[dict[str, object]]],
    *,
    document_title: str,
    section_subtitle: str | None = None,
) -> bytes:
    pages: list[list[str]] = [[document_title, ""]]
    if not grouped:
        pages[0].append("No schedule items found.")
    else:
        pages = []
        for class_name, class_rows in grouped.items():
            page_lines = [document_title, ""] + _grid_to_pdf_lines(
                class_name, class_rows, subtitle=section_subtitle
            )
            pages.append(page_lines)

    return _build_multipage_pdf(pages)


def _build_multipage_pdf(pages: list[list[str]]) -> bytes:
    """Minimal multi-page PDF (Helvetica, 14pt leading)."""
    page_streams: list[bytes] = []
    for lines in pages:
        stream_parts = ["BT", "/F1 10 Tf", "36 806 Td", "14 TL"]
        first = True
        for line in lines:
            safe = _escape_pdf_text(line)
            if first:
                stream_parts.append(f"({safe}) Tj")
                first = False
            else:
                stream_parts.append(f"T* ({safe}) Tj")
        stream_parts.append("ET")
        page_streams.append("\n".join(stream_parts).encode("latin-1", errors="replace"))

    if not page_streams:
        page_streams = [
            b"BT /F1 10 Tf 36 806 Td (Empty export) Tj ET",
        ]

    # Build PDF objects: catalog, pages, page nodes, font, content streams
    font_obj = b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>"
    page_obj_nums: list[int] = []
    content_obj_nums: list[int] = []
    # 1 catalog, 2 pages, 3 font, then pairs of page+content
    next_num = 4
    page_entries: list[bytes] = []
    content_objects: list[bytes] = []

    for stream in page_streams:
        page_obj_nums.append(next_num)
        next_num += 1
        content_obj_nums.append(next_num)
        next_num += 1
        page_entries.append(
            f"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] "
            f"/Resources << /Font << /F1 3 0 R >> >> /Contents {content_obj_nums[-1]} 0 R >>".encode("latin-1")
        )
        content_objects.append(
            f"<< /Length {len(stream)} >>\nstream\n".encode("latin-1") + stream + b"\nendstream"
        )

    kids = " ".join(f"{n} 0 R" for n in page_obj_nums)
    pages_obj = f"<< /Type /Pages /Kids [{kids}] /Count {len(page_obj_nums)} >>".encode("latin-1")
    catalog_obj = b"<< /Type /Catalog /Pages 2 0 R >>"

    objects: list[bytes] = [catalog_obj, pages_obj, font_obj]
    for page_obj, content_obj in zip(page_entries, content_objects, strict=True):
        objects.append(page_obj)
        objects.append(content_obj)

    out = bytearray(b"%PDF-1.4\n")
    offsets: list[int] = [0]
    for idx, obj in enumerate(objects, start=1):
        offsets.append(len(out))
        out.extend(f"{idx} 0 obj\n".encode("latin-1"))
        out.extend(obj)
        out.extend(b"\nendobj\n")
    xref_offset = len(out)
    out.extend(f"xref\n0 {len(objects) + 1}\n".encode("latin-1"))
    out.extend(b"0000000000 65535 f \n")
    for off in offsets[1:]:
        out.extend(f"{off:010d} 00000 n \n".encode("latin-1"))
    out.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n{xref_offset}\n%%EOF\n".encode("latin-1")
    )
    return bytes(out)


def build_schedule_export(
    db: Session,
    school_id: int,
    *,
    view: ExportView,
    fmt: ExportFormat,
    entity_id: int | None = None,
    simple: bool = False,
) -> tuple[bytes, str, str]:
    if simple:
        view = "school"
        entity_id = None
    school = db.get(School, school_id)
    school_label = school.name if school else f"School {school_id}"

    rows = _collect_rows(db, school_id, view, entity_id)
    grouped = _group_rows_by_class(rows)

    if view == "class" and entity_id is not None:
        class_row = db.get(StudentClass, entity_id)
        class_label = class_row.class_name if class_row else "class"
        document_title = f"{school_label} — Class {class_label}"
        filename_stem = f"schedule_{class_label}"
        layout: Literal["multi_sheet", "sections"] = "sections"
        section_subtitle = None
    elif view == "teacher" and entity_id is not None:
        teacher = db.get(Teacher, entity_id)
        teacher_label = teacher.full_name if teacher else f"teacher_{entity_id}"
        document_title = f"{school_label} — Teacher {teacher_label}"
        filename_stem = f"schedule_teacher"
        layout = "sections"
        section_subtitle = f"Teacher: {teacher_label}"
    else:
        document_title = f"{school_label} — Full school timetable"
        filename_stem = "schedule_school"
        layout = "multi_sheet"
        section_subtitle = None

    if fmt == "xlsx":
        payload = _build_xlsx(
            grouped,
            document_title=document_title,
            layout=layout,
            section_subtitle=section_subtitle,
        )
        return payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", f"{filename_stem}.xlsx"

    payload = _build_pdf(
        grouped,
        document_title=document_title,
        section_subtitle=section_subtitle,
    )
    return payload, "application/pdf", f"{filename_stem}.pdf"
