"""Excel data import: workbook template, parsing, validation and commit.

Design notes:
- One workbook per school. The user downloads a template via GET /imports/template,
  fills natural-key cells (no DB ids), and uploads back.
- Validate is read-only: it builds an ImportPlan that knows which rows already
  exist in the DB (matched by natural key inside the school) and which are new.
- Commit applies the plan inside a single SQLAlchemy transaction, honouring
  per-sheet modes. If anything raises, the transaction is rolled back.
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from datetime import time
from typing import Any, Callable

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models.entities import (
    ClassSubjectHours,
    Classroom,
    ClassroomSpecialization,
    GroupFlow,
    LessonSlot,
    ScheduleItem,
    StudentClass,
    Subject,
    Teacher,
)
from app.schemas.imports import (
    ALL_SHEETS,
    CommitSheetResult,
    ImportIssue,
    ImportMode,
    IssueSeverity,
    SHEET_CLASSES,
    SHEET_CLASSROOMS,
    SHEET_CURRICULUM,
    SHEET_GROUP_FLOWS,
    SHEET_LESSON_SLOTS,
    SHEET_SCHEDULE,
    SHEET_SUBJECTS,
    SHEET_TEACHERS,
    SheetStats,
    allowed_modes_for,
    default_modes,
)


SHEET_COLUMNS: dict[str, list[tuple[str, str]]] = {
    SHEET_SUBJECTS: [
        ("name", "Название предмета (уникально) · Subject name (unique)"),
        ("requires_special_room", "true / false — нужен спецкабинет? · needs special room?"),
        ("required_specialization", "Если нужен: standard, chemistry_lab, physics_lab, gym, language_room"),
    ],
    SHEET_LESSON_SLOTS: [
        ("day_of_week", "1=Пн … 7=Вс · 1=Mon … 7=Sun"),
        ("lesson_number", "Номер урока в этот день · lesson index in the day"),
        ("start_time", "Начало ЧЧ:ММ · start HH:MM"),
        ("end_time", "Конец ЧЧ:ММ · end HH:MM"),
    ],
    SHEET_CLASSES: [
        ("class_name", "Название параллели, уникально в школе · class label, unique per school"),
        ("students_count", "Количество учеников · headcount"),
    ],
    SHEET_TEACHERS: [
        ("full_name", "ФИО, уникально в школе · full name, unique per school"),
        ("subject_1", "Предмет 1 — выберите из списка (лист Subjects) · subject 1 from dropdown"),
        ("subject_2", "Предмет 2 (необязательно) · optional subject 2"),
        ("subject_3", "Предмет 3 (необязательно) · optional subject 3"),
        ("subject_4", "Предмет 4 (необязательно) · optional subject 4"),
        ("subject_5", "Предмет 5 (необязательно) · optional subject 5"),
        ("weekly_load_limit", "Макс. часов в неделю, 0 = без лимита · max hours/week, 0 = no cap"),
        ("unavailable_day_1", "Недоступный день 1..7 (необяз.) · optional weekday 1=Mon … 7=Sun"),
        ("unavailable_day_2", "Второй недоступный день · optional second weekday"),
        ("unavailable_day_3", "Третий недоступный день · optional third weekday"),
    ],
    SHEET_CLASSROOMS: [
        ("room_number", "Номер/название кабинета · room label, unique per school"),
        ("capacity", "Вместимость (учеников) · seat capacity"),
        ("specialization", "Тип: standard, chemistry_lab, physics_lab, gym, language_room"),
    ],
    SHEET_GROUP_FLOWS: [
        ("group_name", "Название потока · flow name, unique per school"),
        ("class_1", "Класс 1 — выберите из списка (лист Classes) · class 1 from dropdown"),
        ("class_2", "Класс 2 (необязательно) · optional class 2"),
        ("class_3", "Класс 3 (необязательно) · optional class 3"),
        ("class_4", "Класс 4 (необязательно) · optional class 4"),
        ("class_5", "Класс 5 (необязательно) · optional class 5"),
    ],
    SHEET_CURRICULUM: [
        ("class_name", "Класс из листа Classes · class from Classes sheet"),
        ("subject_name", "Предмет из Subjects · subject from Subjects sheet"),
        ("hours_per_week", "Часов в неделю по этому предмету · planned hours per week"),
    ],
    SHEET_SCHEDULE: [
        ("class_name", "Класс · class"),
        ("subject_name", "Предмет · subject"),
        ("teacher_full_name", "Учитель (как в Teachers) · teacher full name"),
        ("room_number", "Кабинет (как в Classrooms) · room number"),
        ("day_of_week", "День 1..7 · weekday"),
        ("lesson_number", "Номер слота в день · slot index in day"),
        ("is_grouped", "true / false — групповой урок? · grouped lesson?"),
        ("group_name", "Если grouped: имя потока (GroupFlows) · flow name if grouped"),
    ],
}


SPECIALIZATIONS = {member.value for member in ClassroomSpecialization}

# Data-entry grid: rows 4..TEMPLATE_LAST_ROW (row 3 is the green banner).
_TEMPLATE_FIRST_ROW = 4
_TEMPLATE_LAST_ROW = 503

_WEEKDAY_LIST = '"1,2,3,4,5,6,7"'
_LESSON_NUMBER_LIST = '"1,2,3,4,5,6,7,8,9,10,11,12"'

# Named ranges for cross-sheet dropdowns (defined in build_template_workbook).
_NAME_SUBJECTS = "AtlasSubjects"
_NAME_CLASSES = "AtlasClasses"
_NAME_TEACHERS = "AtlasTeachers"
_NAME_CLASSROOMS = "AtlasClassrooms"
_NAME_GROUP_FLOWS = "AtlasGroupFlows"

# Template / workbook styling (hex without FF prefix for openpyxl PatternFill)
_FILL_HEADER = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FILL_HINT = PatternFill(start_color="D6E4F4", end_color="D6E4F4", fill_type="solid")
_FILL_BANNER = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
_FILL_README_TITLE = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_FILL_README_SECTION = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
_FILL_README_TABLE_HEAD = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
_FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
_FONT_HINT = Font(name="Calibri", size=9, italic=True, color="1F3864")
_FONT_BODY = Font(name="Calibri", size=11, color="000000")
_FONT_README_TITLE = Font(name="Calibri", size=18, bold=True, color="FFFFFF")
_FONT_README_SECTION = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
_FONT_README_SMALL = Font(name="Calibri", size=10, color="000000")
_BORDER_THIN = Border(
    left=Side(style="thin", color="8FAADC"),
    right=Side(style="thin", color="8FAADC"),
    top=Side(style="thin", color="8FAADC"),
    bottom=Side(style="thin", color="8FAADC"),
)
_ALIGN_HEADER = Alignment(horizontal="center", vertical="center", wrap_text=True)
_ALIGN_LEFT_WRAP = Alignment(vertical="top", wrap_text=True)


def _all_nonempty_cells_are_comment_hints(cells: list[Any]) -> bool:
    """True for template hint rows: every non-empty cell is a comment (# …)."""

    seen = False
    for cell in cells:
        if cell is None:
            continue
        text = str(cell).strip()
        if not text:
            continue
        seen = True
        if not (text.startswith("# ") or text == "#"):
            return False
    return seen


# ---------------------------------------------------------------------------
# Plan dataclasses
# ---------------------------------------------------------------------------


@dataclass
class _RowOp:
    """One parsed row with its outcome (create / update / skip)."""

    row_number: int
    action: str  # "create" | "update" | "skip" | "error"
    payload: dict[str, Any] = field(default_factory=dict)
    existing_id: int | None = None
    natural_key: tuple[Any, ...] | None = None


@dataclass
class SheetPlan:
    sheet: str
    rows_total: int = 0
    operations: list[_RowOp] = field(default_factory=list)
    rows_with_errors: int = 0


@dataclass
class ImportPlan:
    school_id: int
    sheets: dict[str, SheetPlan] = field(default_factory=dict)
    issues: list[ImportIssue] = field(default_factory=list)

    @property
    def error_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.error.value)

    @property
    def warning_count(self) -> int:
        return sum(1 for issue in self.issues if issue.severity == IssueSeverity.warning.value)


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------


def _coerce_str(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _coerce_int(value: Any) -> int | None:
    if value is None or value == "":
        return None
    if isinstance(value, bool):
        return int(value)
    if isinstance(value, int):
        return int(value)
    if isinstance(value, float):
        if value.is_integer():
            return int(value)
        return None
    text = str(value).strip()
    if not text:
        return None
    try:
        return int(text)
    except ValueError:
        try:
            f = float(text)
        except ValueError:
            return None
        if f.is_integer():
            return int(f)
        return None


def _coerce_bool(value: Any) -> bool:
    if value is None or value == "":
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return bool(value)
    text = str(value).strip().lower()
    return text in {"1", "true", "yes", "y", "да", "иә"}


def _coerce_time(value: Any) -> time | None:
    if value is None or value == "":
        return None
    if isinstance(value, time):
        return value
    text = str(value).strip()
    if not text:
        return None
    parts = text.split(":")
    try:
        if len(parts) == 2:
            return time(int(parts[0]), int(parts[1]))
        if len(parts) >= 3:
            return time(int(parts[0]), int(parts[1]), int(parts[2]))
    except ValueError:
        return None
    return None


def _split_csv(value: Any) -> list[str]:
    text = _coerce_str(value)
    if not text:
        return []
    return [piece.strip() for piece in text.split(",") if piece.strip()]


# ---------------------------------------------------------------------------
# Workbook reading
# ---------------------------------------------------------------------------


def _read_sheet_rows(workbook: Workbook, sheet_name: str) -> tuple[list[str], list[tuple[int, list[Any]]]]:
    """Return (header_columns_lowercased, rows_with_excel_row_numbers).

    Rows that are completely empty are filtered out. Excel row numbers start at 1
    (so the first data row is 2).
    """

    if sheet_name not in workbook.sheetnames:
        return [], []
    sheet = workbook[sheet_name]
    rows_iter = sheet.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        return [], []
    headers = [str(value).strip().lower() if value is not None else "" for value in header_row]
    rows: list[tuple[int, list[Any]]] = []
    for excel_index, raw_row in enumerate(rows_iter, start=2):
        if raw_row is None:
            continue
        cells = list(raw_row)
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in cells):
            continue
        if _all_nonempty_cells_are_comment_hints(cells):
            continue
        rows.append((excel_index, cells))
    return headers, rows


def _index_columns(headers: list[str], expected: list[tuple[str, str]], *, legacy: tuple[str, ...] = ()) -> dict[str, int]:
    """Map expected column name -> position in the row (or missing)."""

    mapping: dict[str, int] = {}
    for column, _ in expected:
        try:
            mapping[column] = headers.index(column.lower())
        except ValueError:
            mapping[column] = -1
    for column in legacy:
        if mapping.get(column, -1) >= 0:
            continue
        try:
            mapping[column] = headers.index(column.lower())
        except ValueError:
            mapping[column] = -1
    return mapping


def _cell(row: list[Any], indexes: dict[str, int], column: str) -> Any:
    pos = indexes.get(column, -1)
    if pos < 0 or pos >= len(row):
        return None
    return row[pos]


# ---------------------------------------------------------------------------
# Template generation
# ---------------------------------------------------------------------------


def _sheet_title_and_subtitle(sheet_name: str) -> tuple[str, str]:
    """RU + EN titles for the green banner row (must start with '# ' for import skip)."""

    labels: dict[str, tuple[str, str]] = {
        SHEET_SUBJECTS: ("Предметы", "Subjects — глобальный справочник"),
        SHEET_LESSON_SLOTS: ("Слоты уроков", "Lesson slots — время сетки"),
        SHEET_CLASSES: ("Классы", "Classes — параллели школы"),
        SHEET_TEACHERS: ("Учителя", "Teachers"),
        SHEET_CLASSROOMS: ("Кабинеты", "Classrooms"),
        SHEET_GROUP_FLOWS: ("Потоки", "Group flows — объединённые классы"),
        SHEET_CURRICULUM: ("Учебный план", "Curriculum — часы в неделю"),
        SHEET_SCHEDULE: ("Расписание", "Schedule"),
    }
    ru, en = labels.get(sheet_name, (sheet_name, ""))
    return ru, en


def _column_letter(sheet_name: str, column_name: str) -> str:
    for index, (name, _hint) in enumerate(SHEET_COLUMNS[sheet_name], start=1):
        if name == column_name:
            return get_column_letter(index)
    raise KeyError(f"Column {column_name!r} not on sheet {sheet_name!r}")


def _sheet_column_range(sheet_name: str, column_name: str, first_row: int = _TEMPLATE_FIRST_ROW, last_row: int = _TEMPLATE_LAST_ROW) -> str:
    col = _column_letter(sheet_name, column_name)
    return f"'{sheet_name}'!${col}${first_row}:${col}${last_row}"


def _define_list_name(wb: Workbook, name: str, range_ref: str) -> None:
    wb.defined_names.add(DefinedName(name, attr_text=range_ref))


def _merge_csv_and_slots(
    row: list[Any],
    indexes: dict[str, int],
    legacy_column: str,
    slot_columns: tuple[str, ...],
) -> list[str]:
    values = _split_csv(_cell(row, indexes, legacy_column))
    for column in slot_columns:
        pos = indexes.get(column, -1)
        if pos < 0:
            continue
        item = _coerce_str(_cell(row, indexes, column))
        if item:
            values.append(item)
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        if value in seen:
            continue
        seen.add(value)
        ordered.append(value)
    return ordered


def _apply_list_validation(
    ws,
    col_letter: str,
    formula1: str,
    *,
    first_row: int = _TEMPLATE_FIRST_ROW,
    last_row: int = _TEMPLATE_LAST_ROW,
    allow_blank: bool = True,
) -> None:
    dv = DataValidation(type="list", formula1=formula1, allow_blank=allow_blank)
    dv.error = "Выберите значение из списка (заполните листы выше по порядку)."
    dv.errorTitle = "Неверное значение"
    dv.errorStyle = "stop"
    dv.showDropDown = True
    ws.add_data_validation(dv)
    dv.add(f"{col_letter}{first_row}:{col_letter}{last_row}")


def _style_data_sheet(ws, sheet_name: str, columns: list[tuple[str, str]]) -> None:
    num_cols = len(columns)
    last_letter = get_column_letter(num_cols)
    ru_title, en_sub = _sheet_title_and_subtitle(sheet_name)

    for col_index, (name, hint) in enumerate(columns, start=1):
        letter = get_column_letter(col_index)
        cell_h = ws.cell(row=1, column=col_index, value=name)
        cell_h.font = _FONT_HEADER
        cell_h.fill = _FILL_HEADER
        cell_h.alignment = _ALIGN_HEADER
        cell_h.border = _BORDER_THIN

        cell_hint = ws.cell(row=2, column=col_index, value=f"# {hint}")
        cell_hint.font = _FONT_HINT
        cell_hint.fill = _FILL_HINT
        cell_hint.alignment = Alignment(vertical="top", wrap_text=True)
        cell_hint.border = _BORDER_THIN

        width = max(14, min(48, len(name) + len(hint) // 6 + 6))
        ws.column_dimensions[letter].width = width

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = max(28, 14 + 4 * (num_cols // 4))

    banner = (
        f"# ↓ {ru_title} · {en_sub} — вводите данные с 4-й строки. "
        f"Rows starting with «# » are ignored. Delete this row if you want."
    )
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=num_cols)
    top_left = ws.cell(row=3, column=1, value=banner)
    top_left.font = Font(name="Calibri", size=10, bold=True, color="375623")
    top_left.fill = _FILL_BANNER
    top_left.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    top_left.border = _BORDER_THIN
    for c in range(2, num_cols + 1):
        ws.cell(row=3, column=c).border = _BORDER_THIN

    grid_bottom = 42
    for r in range(4, grid_bottom + 1):
        for c in range(1, num_cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = _BORDER_THIN
            cell.fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
            cell.font = _FONT_BODY

    ws.freeze_panes = "A4"
    ws.auto_filter.ref = f"A1:{last_letter}5000"
    ws.sheet_properties.tabColor = Color(rgb="FF1F4E79")

    spec_options = ",".join(sorted(SPECIALIZATIONS))
    bool_formula = '"true,false"'

    if sheet_name == SHEET_SUBJECTS:
        _apply_list_validation(ws, _column_letter(sheet_name, "requires_special_room"), bool_formula)
        _apply_list_validation(ws, _column_letter(sheet_name, "required_specialization"), f'"{spec_options}"')
    elif sheet_name == SHEET_LESSON_SLOTS:
        _apply_list_validation(ws, _column_letter(sheet_name, "day_of_week"), _WEEKDAY_LIST)
        _apply_list_validation(ws, _column_letter(sheet_name, "lesson_number"), _LESSON_NUMBER_LIST)
    elif sheet_name == SHEET_TEACHERS:
        subj_ref = f"={_NAME_SUBJECTS}"
        for col in ("subject_1", "subject_2", "subject_3", "subject_4", "subject_5"):
            _apply_list_validation(ws, _column_letter(sheet_name, col), subj_ref)
        for col in ("unavailable_day_1", "unavailable_day_2", "unavailable_day_3"):
            _apply_list_validation(ws, _column_letter(sheet_name, col), _WEEKDAY_LIST)
    elif sheet_name == SHEET_CLASSROOMS:
        _apply_list_validation(ws, _column_letter(sheet_name, "specialization"), f'"{spec_options}"')
    elif sheet_name == SHEET_GROUP_FLOWS:
        class_ref = f"={_NAME_CLASSES}"
        for col in ("class_1", "class_2", "class_3", "class_4", "class_5"):
            _apply_list_validation(ws, _column_letter(sheet_name, col), class_ref)
    elif sheet_name == SHEET_CURRICULUM:
        _apply_list_validation(ws, _column_letter(sheet_name, "class_name"), f"={_NAME_CLASSES}")
        _apply_list_validation(ws, _column_letter(sheet_name, "subject_name"), f"={_NAME_SUBJECTS}")
    elif sheet_name == SHEET_SCHEDULE:
        _apply_list_validation(ws, _column_letter(sheet_name, "class_name"), f"={_NAME_CLASSES}")
        _apply_list_validation(ws, _column_letter(sheet_name, "subject_name"), f"={_NAME_SUBJECTS}")
        _apply_list_validation(ws, _column_letter(sheet_name, "teacher_full_name"), f"={_NAME_TEACHERS}")
        _apply_list_validation(ws, _column_letter(sheet_name, "room_number"), f"={_NAME_CLASSROOMS}")
        _apply_list_validation(ws, _column_letter(sheet_name, "day_of_week"), _WEEKDAY_LIST)
        _apply_list_validation(ws, _column_letter(sheet_name, "lesson_number"), _LESSON_NUMBER_LIST)
        _apply_list_validation(ws, _column_letter(sheet_name, "is_grouped"), bool_formula)
        _apply_list_validation(ws, _column_letter(sheet_name, "group_name"), f"={_NAME_GROUP_FLOWS}")


def _apply_workbook_list_names(wb: Workbook) -> None:
    """Named ranges pointing at natural-key columns — used by dependent-sheet dropdowns."""

    _define_list_name(wb, _NAME_SUBJECTS, _sheet_column_range(SHEET_SUBJECTS, "name"))
    _define_list_name(wb, _NAME_CLASSES, _sheet_column_range(SHEET_CLASSES, "class_name"))
    _define_list_name(wb, _NAME_TEACHERS, _sheet_column_range(SHEET_TEACHERS, "full_name"))
    _define_list_name(wb, _NAME_CLASSROOMS, _sheet_column_range(SHEET_CLASSROOMS, "room_number"))
    _define_list_name(wb, _NAME_GROUP_FLOWS, _sheet_column_range(SHEET_GROUP_FLOWS, "group_name"))


def _build_readme_sheet(ws) -> None:
    """First sheet: instructions with clear typography."""

    ws.sheet_properties.tabColor = Color(rgb="FF2F5496")
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 62

    ws.merge_cells("A1:B1")
    t = ws["A1"]
    t.value = "Atlas · импорт данных школы\nSchool data import"
    t.font = _FONT_README_TITLE
    t.fill = _FILL_README_TITLE
    t.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 52

    blocks: list[tuple[str, str]] = [
        (
            "Как пользоваться · Quick start",
            "1) Скачайте шаблон на сайте (страница «Импорт»).\n"
            "2) Заполняйте листы по порядку: Subjects → LessonSlots → Classes → Teachers → "
            "Classrooms → GroupFlows → Curriculum → Schedule.\n"
            "3) Строки с подсказками (начинаются с «# ») и пустые строки не импортируются.\n"
            "4) Загрузите файл → «Проверить» → при отсутствии ошибок «Применить к базе».",
        ),
        (
            "Советы · Tips",
            "• Время: формат ЧЧ:ММ (например 08:30).\n"
            "• Дни недели: 1 = понедельник … 7 = воскресенье.\n"
            "• Заполняйте листы строго по порядку — в следующих листах появятся выпадающие списки.\n"
            "• Предметы, классы, учителя, кабинеты — только из списка (нельзя ввести с опечаткой).\n"
            "• Учитель: до 5 предметов в колонках subject_1…subject_5. Поток: до 5 классов class_1…class_5.",
        ),
        (
            "Листы · Sheets",
            "Каждый лист — отдельная таблица. Синяя строка 1 — имена полей для сервера; "
            "голубая строка 2 — подсказки; зелёная строка 3 — зона ввода с 4-й строки.",
        ),
    ]

    row = 3
    for title, body in blocks:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        head = ws.cell(row=row, column=1, value=title)
        head.font = _FONT_README_SECTION
        head.fill = _FILL_README_SECTION
        head.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        head.border = _BORDER_THIN
        ws.cell(row=row, column=2).border = _BORDER_THIN
        ws.row_dimensions[row].height = 22
        row += 1

        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        body_cell = ws.cell(row=row, column=1, value=body)
        body_cell.font = _FONT_README_SMALL
        body_cell.alignment = _ALIGN_LEFT_WRAP
        body_cell.border = _BORDER_THIN
        ws.cell(row=row, column=2).border = _BORDER_THIN
        ws.row_dimensions[row].height = max(48, 14 * (body.count("\n") + 1))
        row += 1
        row += 1

    # Table of sheets
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
    th = ws.cell(row=row, column=1, value="Состав шаблона · Workbook layout")
    th.font = _FONT_README_SECTION
    th.fill = _FILL_README_SECTION
    th.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.cell(row=row, column=2).border = _BORDER_THIN
    ws.row_dimensions[row].height = 22
    row += 1

    table_headers = ("Лист / Sheet", "Назначение · Purpose")
    for col, text in enumerate(table_headers, start=1):
        c = ws.cell(row=row, column=col, value=text)
        c.font = Font(name="Calibri", size=10, bold=True, color="000000")
        c.fill = _FILL_README_TABLE_HEAD
        c.border = _BORDER_THIN
        c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 20
    row += 1

    for sheet_name in ALL_SHEETS:
        ru, en = _sheet_title_and_subtitle(sheet_name)
        ws.cell(row=row, column=1, value=sheet_name).font = Font(name="Calibri", size=10, bold=True)
        ws.cell(row=row, column=1).border = _BORDER_THIN
        ws.cell(row=row, column=2, value=f"{ru} — {en}").font = _FONT_README_SMALL
        ws.cell(row=row, column=2).border = _BORDER_THIN
        ws.cell(row=row, column=2).alignment = _ALIGN_LEFT_WRAP
        ws.row_dimensions[row].height = 18
        row += 1


def build_template_workbook() -> bytes:
    """Generate a styled .xlsx template: README, column headers, hints, entry grid, validations."""

    wb = Workbook()
    default = wb.active
    wb.remove(default)

    instructions = wb.create_sheet(title="_README", index=0)
    _build_readme_sheet(instructions)

    for sheet_name in ALL_SHEETS:
        ws = wb.create_sheet(title=sheet_name)
        columns = SHEET_COLUMNS[sheet_name]
        for col_index, (name, hint) in enumerate(columns, start=1):
            ws.cell(row=1, column=col_index, value=name)
            ws.cell(row=2, column=col_index, value=f"# {hint}")
        _style_data_sheet(ws, sheet_name, columns)

    _apply_workbook_list_names(wb)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def load_workbook_from_bytes(data: bytes) -> Workbook:
    return load_workbook(io.BytesIO(data), data_only=True)


# ---------------------------------------------------------------------------
# Plan-building (validate)
# ---------------------------------------------------------------------------


@dataclass
class _SchoolContext:
    """Snapshot of existing rows inside this school, keyed by natural key."""

    school_id: int
    subjects: dict[str, Subject] = field(default_factory=dict)
    lesson_slots: dict[tuple[int, int], LessonSlot] = field(default_factory=dict)
    classes: dict[str, StudentClass] = field(default_factory=dict)
    teachers: dict[str, Teacher] = field(default_factory=dict)
    classrooms: dict[str, Classroom] = field(default_factory=dict)
    group_flows: dict[str, GroupFlow] = field(default_factory=dict)
    curriculum: dict[tuple[int, int], ClassSubjectHours] = field(default_factory=dict)
    schedule: dict[tuple[int, int], ScheduleItem] = field(default_factory=dict)
    # Workbook-introduced new natural keys, available to dependent sheets
    pending_classes: set[str] = field(default_factory=set)
    pending_subjects: set[str] = field(default_factory=set)
    pending_teachers: set[str] = field(default_factory=set)
    pending_classrooms: set[str] = field(default_factory=set)
    pending_lesson_slots: set[tuple[int, int]] = field(default_factory=set)
    pending_group_flows: set[str] = field(default_factory=set)


def _build_context(db: Session, school_id: int) -> _SchoolContext:
    ctx = _SchoolContext(school_id=school_id)
    for subj in db.scalars(select(Subject)):
        ctx.subjects[subj.name] = subj
    for slot in db.scalars(select(LessonSlot)):
        ctx.lesson_slots[(slot.day_of_week, slot.lesson_number)] = slot
    for cls in db.scalars(select(StudentClass).where(StudentClass.school_id == school_id)):
        ctx.classes[cls.class_name] = cls
    for tch in db.scalars(select(Teacher).where(Teacher.school_id == school_id)):
        ctx.teachers[tch.full_name] = tch
    for room in db.scalars(select(Classroom).where(Classroom.school_id == school_id)):
        ctx.classrooms[room.room_number] = room
    for flow in db.scalars(select(GroupFlow).where(GroupFlow.school_id == school_id)):
        ctx.group_flows[flow.group_name] = flow
    for plan in db.scalars(select(ClassSubjectHours).where(ClassSubjectHours.school_id == school_id)):
        ctx.curriculum[(plan.class_id, plan.subject_id)] = plan
    for item in db.scalars(select(ScheduleItem).where(ScheduleItem.school_id == school_id)):
        ctx.schedule[(item.class_id, item.lesson_slot_id)] = item
    return ctx


# ---- per-sheet parsers ----------------------------------------------------


def _parse_subjects(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_SUBJECTS, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_SUBJECTS])
    if indexes["name"] < 0:
        issues.append(ImportIssue(sheet=SHEET_SUBJECTS, severity=IssueSeverity.error, code="missing_column", message="Required column 'name' is missing"))
        return plan
    seen: set[str] = set()
    for excel_row, row in rows:
        name = _coerce_str(_cell(row, indexes, "name"))
        if not name:
            issues.append(ImportIssue(sheet=SHEET_SUBJECTS, row=excel_row, column="name", severity=IssueSeverity.error, code="required", message="Subject name is required"))
            plan.rows_with_errors += 1
            continue
        if name in seen:
            issues.append(ImportIssue(sheet=SHEET_SUBJECTS, row=excel_row, column="name", severity=IssueSeverity.error, code="duplicate", message=f"Duplicate subject name '{name}' in workbook"))
            plan.rows_with_errors += 1
            continue
        seen.add(name)
        spec_raw = _coerce_str(_cell(row, indexes, "required_specialization"))
        spec = spec_raw if spec_raw and spec_raw in SPECIALIZATIONS else None
        if spec_raw and spec is None:
            issues.append(ImportIssue(sheet=SHEET_SUBJECTS, row=excel_row, column="required_specialization", severity=IssueSeverity.error, code="invalid_enum", message=f"Unknown specialization '{spec_raw}'"))
            plan.rows_with_errors += 1
            continue
        payload = {
            "name": name,
            "requires_special_room": _coerce_bool(_cell(row, indexes, "requires_special_room")),
            "required_specialization": spec,
        }
        existing = ctx.subjects.get(name)
        action = "update" if existing else "create"
        if action == "create":
            ctx.pending_subjects.add(name)
        plan.operations.append(_RowOp(row_number=excel_row, action=action, payload=payload, existing_id=existing.id if existing else None, natural_key=(name,)))
    return plan


def _parse_lesson_slots(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_LESSON_SLOTS, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_LESSON_SLOTS])
    for required in ("day_of_week", "lesson_number", "start_time", "end_time"):
        if indexes[required] < 0:
            issues.append(ImportIssue(sheet=SHEET_LESSON_SLOTS, severity=IssueSeverity.error, code="missing_column", message=f"Required column '{required}' is missing"))
            return plan
    seen: set[tuple[int, int]] = set()
    for excel_row, row in rows:
        day = _coerce_int(_cell(row, indexes, "day_of_week"))
        lesson = _coerce_int(_cell(row, indexes, "lesson_number"))
        start = _coerce_time(_cell(row, indexes, "start_time"))
        end = _coerce_time(_cell(row, indexes, "end_time"))
        bad = False
        if day is None or not (1 <= day <= 7):
            issues.append(ImportIssue(sheet=SHEET_LESSON_SLOTS, row=excel_row, column="day_of_week", severity=IssueSeverity.error, code="invalid_day", message="day_of_week must be 1..7"))
            bad = True
        if lesson is None or lesson <= 0:
            issues.append(ImportIssue(sheet=SHEET_LESSON_SLOTS, row=excel_row, column="lesson_number", severity=IssueSeverity.error, code="invalid_lesson_number", message="lesson_number must be a positive integer"))
            bad = True
        if start is None:
            issues.append(ImportIssue(sheet=SHEET_LESSON_SLOTS, row=excel_row, column="start_time", severity=IssueSeverity.error, code="invalid_time", message="start_time must be HH:MM"))
            bad = True
        if end is None:
            issues.append(ImportIssue(sheet=SHEET_LESSON_SLOTS, row=excel_row, column="end_time", severity=IssueSeverity.error, code="invalid_time", message="end_time must be HH:MM"))
            bad = True
        if bad:
            plan.rows_with_errors += 1
            continue
        key = (day, lesson)
        if key in seen:
            issues.append(ImportIssue(sheet=SHEET_LESSON_SLOTS, row=excel_row, severity=IssueSeverity.error, code="duplicate", message=f"Duplicate slot day={day} lesson={lesson} in workbook"))
            plan.rows_with_errors += 1
            continue
        seen.add(key)
        existing = ctx.lesson_slots.get(key)
        action = "update" if existing else "create"
        if action == "create":
            ctx.pending_lesson_slots.add(key)
        plan.operations.append(_RowOp(row_number=excel_row, action=action, payload={"day_of_week": day, "lesson_number": lesson, "start_time": start, "end_time": end}, existing_id=existing.id if existing else None, natural_key=key))
    return plan


def _parse_classes(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_CLASSES, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_CLASSES])
    if indexes["class_name"] < 0 or indexes["students_count"] < 0:
        issues.append(ImportIssue(sheet=SHEET_CLASSES, severity=IssueSeverity.error, code="missing_column", message="Required columns: class_name, students_count"))
        return plan
    seen: set[str] = set()
    for excel_row, row in rows:
        name = _coerce_str(_cell(row, indexes, "class_name"))
        count = _coerce_int(_cell(row, indexes, "students_count"))
        if not name:
            issues.append(ImportIssue(sheet=SHEET_CLASSES, row=excel_row, column="class_name", severity=IssueSeverity.error, code="required", message="class_name is required"))
            plan.rows_with_errors += 1
            continue
        if count is None or count < 0:
            issues.append(ImportIssue(sheet=SHEET_CLASSES, row=excel_row, column="students_count", severity=IssueSeverity.error, code="invalid_int", message="students_count must be a non-negative integer"))
            plan.rows_with_errors += 1
            continue
        if name in seen:
            issues.append(ImportIssue(sheet=SHEET_CLASSES, row=excel_row, column="class_name", severity=IssueSeverity.error, code="duplicate", message=f"Duplicate class_name '{name}' in workbook"))
            plan.rows_with_errors += 1
            continue
        seen.add(name)
        existing = ctx.classes.get(name)
        action = "update" if existing else "create"
        if action == "create":
            ctx.pending_classes.add(name)
        plan.operations.append(_RowOp(row_number=excel_row, action=action, payload={"class_name": name, "students_count": count, "school_id": ctx.school_id}, existing_id=existing.id if existing else None, natural_key=(name,)))
    return plan


def _parse_teachers(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_TEACHERS, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_TEACHERS], legacy=("subjects", "unavailable_days"))
    if indexes["full_name"] < 0:
        issues.append(ImportIssue(sheet=SHEET_TEACHERS, severity=IssueSeverity.error, code="missing_column", message="Required column: full_name"))
        return plan
    known_subjects = set(ctx.subjects.keys()) | ctx.pending_subjects
    seen: set[str] = set()
    for excel_row, row in rows:
        full_name = _coerce_str(_cell(row, indexes, "full_name"))
        if not full_name:
            issues.append(ImportIssue(sheet=SHEET_TEACHERS, row=excel_row, column="full_name", severity=IssueSeverity.error, code="required", message="full_name is required"))
            plan.rows_with_errors += 1
            continue
        if full_name in seen:
            issues.append(ImportIssue(sheet=SHEET_TEACHERS, row=excel_row, column="full_name", severity=IssueSeverity.error, code="duplicate", message=f"Duplicate teacher '{full_name}' in workbook"))
            plan.rows_with_errors += 1
            continue
        seen.add(full_name)
        subjects = _merge_csv_and_slots(
            row,
            indexes,
            "subjects",
            ("subject_1", "subject_2", "subject_3", "subject_4", "subject_5"),
        )
        weekly_limit = _coerce_int(_cell(row, indexes, "weekly_load_limit")) or 0
        days_raw = _merge_csv_and_slots(
            row,
            indexes,
            "unavailable_days",
            ("unavailable_day_1", "unavailable_day_2", "unavailable_day_3"),
        )
        unavailable_days: list[int] = []
        subject_bad = False
        for subj in subjects:
            if subj not in known_subjects:
                issues.append(
                    ImportIssue(
                        sheet=SHEET_TEACHERS,
                        row=excel_row,
                        column="subject_1",
                        severity=IssueSeverity.error,
                        code="unknown_subject",
                        message=f"Subject '{subj}' is not on the Subjects sheet — pick from the dropdown list",
                    )
                )
                subject_bad = True
        for raw in days_raw:
            value = _coerce_int(raw)
            if value is None or not (1 <= value <= 7):
                issues.append(
                    ImportIssue(
                        sheet=SHEET_TEACHERS,
                        row=excel_row,
                        column="unavailable_day_1",
                        severity=IssueSeverity.warning,
                        code="invalid_day",
                        message=f"Skipping invalid day '{raw}'",
                    )
                )
                continue
            unavailable_days.append(value)
        if subject_bad:
            plan.rows_with_errors += 1
            continue
        payload = {
            "full_name": full_name,
            "subjects": subjects,
            "weekly_load_limit": weekly_limit,
            "unavailable_days": unavailable_days,
            "school_id": ctx.school_id,
        }
        existing = ctx.teachers.get(full_name)
        action = "update" if existing else "create"
        if action == "create":
            ctx.pending_teachers.add(full_name)
        plan.operations.append(_RowOp(row_number=excel_row, action=action, payload=payload, existing_id=existing.id if existing else None, natural_key=(full_name,)))
    return plan


def _parse_classrooms(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_CLASSROOMS, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_CLASSROOMS])
    if indexes["room_number"] < 0 or indexes["capacity"] < 0:
        issues.append(ImportIssue(sheet=SHEET_CLASSROOMS, severity=IssueSeverity.error, code="missing_column", message="Required columns: room_number, capacity"))
        return plan
    seen: set[str] = set()
    for excel_row, row in rows:
        room = _coerce_str(_cell(row, indexes, "room_number"))
        capacity = _coerce_int(_cell(row, indexes, "capacity"))
        if not room:
            issues.append(ImportIssue(sheet=SHEET_CLASSROOMS, row=excel_row, column="room_number", severity=IssueSeverity.error, code="required", message="room_number is required"))
            plan.rows_with_errors += 1
            continue
        if capacity is None or capacity < 0:
            issues.append(ImportIssue(sheet=SHEET_CLASSROOMS, row=excel_row, column="capacity", severity=IssueSeverity.error, code="invalid_int", message="capacity must be a non-negative integer"))
            plan.rows_with_errors += 1
            continue
        if room in seen:
            issues.append(ImportIssue(sheet=SHEET_CLASSROOMS, row=excel_row, column="room_number", severity=IssueSeverity.error, code="duplicate", message=f"Duplicate room_number '{room}' in workbook"))
            plan.rows_with_errors += 1
            continue
        seen.add(room)
        spec_raw = _coerce_str(_cell(row, indexes, "specialization")) or ClassroomSpecialization.standard.value
        if spec_raw not in SPECIALIZATIONS:
            issues.append(ImportIssue(sheet=SHEET_CLASSROOMS, row=excel_row, column="specialization", severity=IssueSeverity.error, code="invalid_enum", message=f"Unknown specialization '{spec_raw}'"))
            plan.rows_with_errors += 1
            continue
        payload = {"room_number": room, "capacity": capacity, "specialization": spec_raw, "school_id": ctx.school_id}
        existing = ctx.classrooms.get(room)
        action = "update" if existing else "create"
        if action == "create":
            ctx.pending_classrooms.add(room)
        plan.operations.append(_RowOp(row_number=excel_row, action=action, payload=payload, existing_id=existing.id if existing else None, natural_key=(room,)))
    return plan


def _parse_group_flows(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_GROUP_FLOWS, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_GROUP_FLOWS], legacy=("combined_classes",))
    has_class_slots = indexes.get("class_1", -1) >= 0
    has_legacy_classes = indexes.get("combined_classes", -1) >= 0
    if indexes["group_name"] < 0 or (not has_class_slots and not has_legacy_classes):
        issues.append(
            ImportIssue(
                sheet=SHEET_GROUP_FLOWS,
                severity=IssueSeverity.error,
                code="missing_column",
                message="Required columns: group_name and class_1 (or legacy combined_classes)",
            )
        )
        return plan
    known_classes = set(ctx.classes.keys()) | ctx.pending_classes
    seen: set[str] = set()
    for excel_row, row in rows:
        name = _coerce_str(_cell(row, indexes, "group_name"))
        members = _merge_csv_and_slots(
            row,
            indexes,
            "combined_classes",
            ("class_1", "class_2", "class_3", "class_4", "class_5"),
        )
        if not name:
            issues.append(ImportIssue(sheet=SHEET_GROUP_FLOWS, row=excel_row, column="group_name", severity=IssueSeverity.error, code="required", message="group_name is required"))
            plan.rows_with_errors += 1
            continue
        if not members:
            issues.append(ImportIssue(sheet=SHEET_GROUP_FLOWS, row=excel_row, column="class_1", severity=IssueSeverity.error, code="required", message="At least one class is required (class_1)"))
            plan.rows_with_errors += 1
            continue
        if name in seen:
            issues.append(ImportIssue(sheet=SHEET_GROUP_FLOWS, row=excel_row, column="group_name", severity=IssueSeverity.error, code="duplicate", message=f"Duplicate group_name '{name}' in workbook"))
            plan.rows_with_errors += 1
            continue
        seen.add(name)
        bad = False
        for member in members:
            if member not in known_classes:
                issues.append(
                    ImportIssue(
                        sheet=SHEET_GROUP_FLOWS,
                        row=excel_row,
                        column="class_1",
                        severity=IssueSeverity.error,
                        code="unknown_class",
                        message=f"Class '{member}' is not on the Classes sheet — pick from the dropdown list",
                    )
                )
                bad = True
        if bad:
            plan.rows_with_errors += 1
            continue
        existing = ctx.group_flows.get(name)
        action = "update" if existing else "create"
        if action == "create":
            ctx.pending_group_flows.add(name)
        plan.operations.append(_RowOp(row_number=excel_row, action=action, payload={"group_name": name, "combined_classes": members, "school_id": ctx.school_id}, existing_id=existing.id if existing else None, natural_key=(name,)))
    return plan


def _parse_curriculum(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_CURRICULUM, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_CURRICULUM])
    for required in ("class_name", "subject_name", "hours_per_week"):
        if indexes[required] < 0:
            issues.append(ImportIssue(sheet=SHEET_CURRICULUM, severity=IssueSeverity.error, code="missing_column", message=f"Required column '{required}' is missing"))
            return plan
    known_classes = set(ctx.classes.keys()) | ctx.pending_classes
    known_subjects = set(ctx.subjects.keys()) | ctx.pending_subjects
    seen: set[tuple[str, str]] = set()
    for excel_row, row in rows:
        class_name = _coerce_str(_cell(row, indexes, "class_name"))
        subject_name = _coerce_str(_cell(row, indexes, "subject_name"))
        hours = _coerce_int(_cell(row, indexes, "hours_per_week"))
        if not class_name or not subject_name:
            issues.append(ImportIssue(sheet=SHEET_CURRICULUM, row=excel_row, severity=IssueSeverity.error, code="required", message="class_name and subject_name are required"))
            plan.rows_with_errors += 1
            continue
        if hours is None or hours <= 0:
            issues.append(ImportIssue(sheet=SHEET_CURRICULUM, row=excel_row, column="hours_per_week", severity=IssueSeverity.error, code="invalid_int", message="hours_per_week must be a positive integer"))
            plan.rows_with_errors += 1
            continue
        if class_name not in known_classes:
            issues.append(ImportIssue(sheet=SHEET_CURRICULUM, row=excel_row, column="class_name", severity=IssueSeverity.error, code="unknown_class", message=f"Class '{class_name}' is not defined for this school"))
            plan.rows_with_errors += 1
            continue
        if subject_name not in known_subjects:
            issues.append(ImportIssue(sheet=SHEET_CURRICULUM, row=excel_row, column="subject_name", severity=IssueSeverity.error, code="unknown_subject", message=f"Subject '{subject_name}' is not defined"))
            plan.rows_with_errors += 1
            continue
        key = (class_name, subject_name)
        if key in seen:
            issues.append(ImportIssue(sheet=SHEET_CURRICULUM, row=excel_row, severity=IssueSeverity.error, code="duplicate", message=f"Duplicate plan row class={class_name} subject={subject_name}"))
            plan.rows_with_errors += 1
            continue
        seen.add(key)
        cls = ctx.classes.get(class_name)
        subj = ctx.subjects.get(subject_name)
        existing = None
        if cls and subj:
            existing = ctx.curriculum.get((cls.id, subj.id))
        action = "update" if existing else "create"
        plan.operations.append(_RowOp(
            row_number=excel_row,
            action=action,
            payload={"class_name": class_name, "subject_name": subject_name, "hours_per_week": hours},
            existing_id=existing.id if existing else None,
            natural_key=key,
        ))
    return plan


def _parse_schedule(ctx: _SchoolContext, rows: list[tuple[int, list[Any]]], headers: list[str], issues: list[ImportIssue]) -> SheetPlan:
    plan = SheetPlan(sheet=SHEET_SCHEDULE, rows_total=len(rows))
    indexes = _index_columns(headers, SHEET_COLUMNS[SHEET_SCHEDULE])
    required = ("class_name", "subject_name", "teacher_full_name", "room_number", "day_of_week", "lesson_number")
    for col in required:
        if indexes[col] < 0:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, severity=IssueSeverity.error, code="missing_column", message=f"Required column '{col}' is missing"))
            return plan
    known_classes = set(ctx.classes.keys()) | ctx.pending_classes
    known_subjects = set(ctx.subjects.keys()) | ctx.pending_subjects
    known_teachers = set(ctx.teachers.keys()) | ctx.pending_teachers
    known_rooms = set(ctx.classrooms.keys()) | ctx.pending_classrooms
    known_slots = set(ctx.lesson_slots.keys()) | ctx.pending_lesson_slots
    known_flows = set(ctx.group_flows.keys()) | ctx.pending_group_flows
    seen: set[tuple[str, int, int]] = set()
    for excel_row, row in rows:
        class_name = _coerce_str(_cell(row, indexes, "class_name"))
        subject_name = _coerce_str(_cell(row, indexes, "subject_name"))
        teacher_name = _coerce_str(_cell(row, indexes, "teacher_full_name"))
        room = _coerce_str(_cell(row, indexes, "room_number"))
        day = _coerce_int(_cell(row, indexes, "day_of_week"))
        lesson = _coerce_int(_cell(row, indexes, "lesson_number"))
        is_grouped = _coerce_bool(_cell(row, indexes, "is_grouped"))
        group_name = _coerce_str(_cell(row, indexes, "group_name"))
        bad = False
        if not class_name or class_name not in known_classes:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, column="class_name", severity=IssueSeverity.error, code="unknown_class", message=f"Class '{class_name}' is not defined"))
            bad = True
        if not subject_name or subject_name not in known_subjects:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, column="subject_name", severity=IssueSeverity.error, code="unknown_subject", message=f"Subject '{subject_name}' is not defined"))
            bad = True
        if not teacher_name or teacher_name not in known_teachers:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, column="teacher_full_name", severity=IssueSeverity.error, code="unknown_teacher", message=f"Teacher '{teacher_name}' is not defined for this school"))
            bad = True
        if not room or room not in known_rooms:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, column="room_number", severity=IssueSeverity.error, code="unknown_room", message=f"Room '{room}' is not defined for this school"))
            bad = True
        if day is None or lesson is None or (day, lesson) not in known_slots:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, severity=IssueSeverity.error, code="unknown_slot", message=f"Slot day={day} lesson={lesson} is not defined"))
            bad = True
        if is_grouped and (not group_name or group_name not in known_flows):
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, column="group_name", severity=IssueSeverity.error, code="unknown_flow", message=f"Group flow '{group_name}' is not defined"))
            bad = True
        if not is_grouped and group_name:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, column="group_name", severity=IssueSeverity.warning, code="group_ignored", message="group_name provided but is_grouped is false; will be ignored"))
            group_name = ""
        if bad:
            plan.rows_with_errors += 1
            continue
        # day/lesson are validated above
        assert day is not None and lesson is not None
        key = (class_name, day, lesson)
        if key in seen:
            issues.append(ImportIssue(sheet=SHEET_SCHEDULE, row=excel_row, severity=IssueSeverity.error, code="duplicate", message=f"Class {class_name} already has another lesson on day={day} lesson={lesson}"))
            plan.rows_with_errors += 1
            continue
        seen.add(key)
        payload = {
            "class_name": class_name,
            "subject_name": subject_name,
            "teacher_full_name": teacher_name,
            "room_number": room,
            "day_of_week": day,
            "lesson_number": lesson,
            "is_grouped": is_grouped,
            "group_name": group_name or None,
        }
        cls = ctx.classes.get(class_name)
        slot = ctx.lesson_slots.get((day, lesson)) if day and lesson else None
        existing = None
        if cls and slot:
            existing = ctx.schedule.get((cls.id, slot.id))
        action = "update" if existing else "create"
        plan.operations.append(_RowOp(row_number=excel_row, action=action, payload=payload, existing_id=existing.id if existing else None, natural_key=key))
    return plan


_PARSERS: dict[str, Callable[..., SheetPlan]] = {
    SHEET_SUBJECTS: _parse_subjects,
    SHEET_LESSON_SLOTS: _parse_lesson_slots,
    SHEET_CLASSES: _parse_classes,
    SHEET_TEACHERS: _parse_teachers,
    SHEET_CLASSROOMS: _parse_classrooms,
    SHEET_GROUP_FLOWS: _parse_group_flows,
    SHEET_CURRICULUM: _parse_curriculum,
    SHEET_SCHEDULE: _parse_schedule,
}


def build_plan(db: Session, workbook: Workbook, school_id: int) -> ImportPlan:
    ctx = _build_context(db, school_id)
    plan = ImportPlan(school_id=school_id)
    # The fixed order below matches dependency direction. Each parser may add to
    # ctx.pending_* sets so downstream sheets can resolve natural keys that are
    # introduced by the workbook itself.
    for sheet_name in ALL_SHEETS:
        if sheet_name not in workbook.sheetnames:
            plan.sheets[sheet_name] = SheetPlan(sheet=sheet_name)
            continue
        headers, rows = _read_sheet_rows(workbook, sheet_name)
        if not headers:
            plan.sheets[sheet_name] = SheetPlan(sheet=sheet_name)
            continue
        parser = _PARSERS[sheet_name]
        sheet_plan = parser(ctx, rows, headers, plan.issues)
        plan.sheets[sheet_name] = sheet_plan
    return plan


def plan_to_summary(plan: ImportPlan, modes: dict[str, ImportMode]) -> list[SheetStats]:
    sheet_stats: list[SheetStats] = []
    for sheet_name in ALL_SHEETS:
        sheet_plan = plan.sheets.get(sheet_name) or SheetPlan(sheet=sheet_name)
        mode = modes.get(sheet_name, default_modes()[sheet_name])
        rows_to_create = sum(1 for op in sheet_plan.operations if op.action == "create")
        rows_to_update = sum(1 for op in sheet_plan.operations if op.action == "update")
        rows_to_skip = 0
        rows_to_replace = 0
        if mode == ImportMode.skip:
            rows_to_skip = sheet_plan.rows_total
            rows_to_create = 0
            rows_to_update = 0
        elif mode == ImportMode.append:
            rows_to_update = 0
        elif mode == ImportMode.replace:
            rows_to_replace = rows_to_create + rows_to_update
            rows_to_create = 0
            rows_to_update = 0
        sheet_stats.append(SheetStats(
            sheet=sheet_name,
            rows_total=sheet_plan.rows_total,
            rows_to_create=rows_to_create,
            rows_to_update=rows_to_update,
            rows_to_replace=rows_to_replace,
            rows_to_skip=rows_to_skip,
            rows_with_errors=sheet_plan.rows_with_errors,
            default_mode=mode,
            allowed_modes=allowed_modes_for(sheet_name),
        ))
    return sheet_stats


# ---------------------------------------------------------------------------
# Commit (transactional apply)
# ---------------------------------------------------------------------------


def apply_plan(
    db: Session,
    plan: ImportPlan,
    modes: dict[str, ImportMode],
) -> list[CommitSheetResult]:
    """Apply the import plan to the database.

    The caller controls the transaction boundary (commit on success, rollback on
    error). Inside this function we use flush() to materialize ids needed by
    later sheets.
    """

    results: list[CommitSheetResult] = []
    school_id = plan.school_id

    # Track the latest natural-key -> ORM object map across sheets so dependent
    # sheets can reference rows created earlier in this commit.
    subjects: dict[str, Subject] = {row.name: row for row in db.scalars(select(Subject))}
    lesson_slots: dict[tuple[int, int], LessonSlot] = {
        (row.day_of_week, row.lesson_number): row for row in db.scalars(select(LessonSlot))
    }
    classes: dict[str, StudentClass] = {row.class_name: row for row in db.scalars(select(StudentClass).where(StudentClass.school_id == school_id))}
    teachers: dict[str, Teacher] = {row.full_name: row for row in db.scalars(select(Teacher).where(Teacher.school_id == school_id))}
    classrooms: dict[str, Classroom] = {row.room_number: row for row in db.scalars(select(Classroom).where(Classroom.school_id == school_id))}
    group_flows: dict[str, GroupFlow] = {row.group_name: row for row in db.scalars(select(GroupFlow).where(GroupFlow.school_id == school_id))}

    for sheet_name in ALL_SHEETS:
        sheet_plan = plan.sheets.get(sheet_name) or SheetPlan(sheet=sheet_name)
        mode = modes.get(sheet_name, default_modes()[sheet_name])
        result = CommitSheetResult(sheet=sheet_name, mode=mode)
        # 'replace' must always run so it can clear existing rows even when the
        # workbook sheet is empty (user intentionally wipes a section). Other
        # modes are no-ops when there are no parsed operations.
        if mode == ImportMode.skip:
            results.append(result)
            continue
        if not sheet_plan.operations and mode != ImportMode.replace:
            results.append(result)
            continue

        if sheet_name == SHEET_SUBJECTS:
            _apply_subjects(db, sheet_plan, mode, subjects, result)
        elif sheet_name == SHEET_LESSON_SLOTS:
            _apply_lesson_slots(db, sheet_plan, mode, lesson_slots, result)
        elif sheet_name == SHEET_CLASSES:
            _apply_classes(db, school_id, sheet_plan, mode, classes, result)
        elif sheet_name == SHEET_TEACHERS:
            _apply_teachers(db, school_id, sheet_plan, mode, teachers, result)
        elif sheet_name == SHEET_CLASSROOMS:
            _apply_classrooms(db, school_id, sheet_plan, mode, classrooms, result)
        elif sheet_name == SHEET_GROUP_FLOWS:
            _apply_group_flows(db, school_id, sheet_plan, mode, classes, group_flows, result)
        elif sheet_name == SHEET_CURRICULUM:
            _apply_curriculum(db, school_id, sheet_plan, mode, classes, subjects, result)
        elif sheet_name == SHEET_SCHEDULE:
            _apply_schedule(db, school_id, sheet_plan, mode, classes, subjects, teachers, classrooms, lesson_slots, group_flows, result)
        results.append(result)
        db.flush()
    return results


def _apply_subjects(db: Session, sheet_plan: SheetPlan, mode: ImportMode, subjects: dict[str, Subject], result: CommitSheetResult) -> None:
    if mode == ImportMode.replace:
        for row in db.scalars(select(Subject)):
            db.delete(row)
        subjects.clear()
        db.flush()
        for op in sheet_plan.operations:
            new = Subject(name=op.payload["name"], requires_special_room=op.payload["requires_special_room"], required_specialization=op.payload["required_specialization"])
            db.add(new)
            db.flush()
            subjects[new.name] = new
            result.created += 1
        return
    for op in sheet_plan.operations:
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(Subject, op.existing_id)
            if row is not None:
                row.requires_special_room = op.payload["requires_special_room"]
                row.required_specialization = op.payload["required_specialization"]
                subjects[row.name] = row
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        new = Subject(name=op.payload["name"], requires_special_room=op.payload["requires_special_room"], required_specialization=op.payload["required_specialization"])
        db.add(new)
        db.flush()
        subjects[new.name] = new
        result.created += 1


def _apply_lesson_slots(db: Session, sheet_plan: SheetPlan, mode: ImportMode, lesson_slots: dict[tuple[int, int], LessonSlot], result: CommitSheetResult) -> None:
    if mode == ImportMode.replace:
        for row in db.scalars(select(LessonSlot)):
            db.delete(row)
        lesson_slots.clear()
        db.flush()
        for op in sheet_plan.operations:
            new = LessonSlot(day_of_week=op.payload["day_of_week"], lesson_number=op.payload["lesson_number"], start_time=op.payload["start_time"], end_time=op.payload["end_time"])
            db.add(new)
            db.flush()
            lesson_slots[(new.day_of_week, new.lesson_number)] = new
            result.created += 1
        return
    for op in sheet_plan.operations:
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(LessonSlot, op.existing_id)
            if row is not None:
                row.start_time = op.payload["start_time"]
                row.end_time = op.payload["end_time"]
                lesson_slots[(row.day_of_week, row.lesson_number)] = row
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        new = LessonSlot(day_of_week=op.payload["day_of_week"], lesson_number=op.payload["lesson_number"], start_time=op.payload["start_time"], end_time=op.payload["end_time"])
        db.add(new)
        db.flush()
        lesson_slots[(new.day_of_week, new.lesson_number)] = new
        result.created += 1


def _apply_classes(db: Session, school_id: int, sheet_plan: SheetPlan, mode: ImportMode, classes: dict[str, StudentClass], result: CommitSheetResult) -> None:
    if mode == ImportMode.replace:
        for row in db.scalars(select(StudentClass).where(StudentClass.school_id == school_id)):
            db.delete(row)
        classes.clear()
        db.flush()
        for op in sheet_plan.operations:
            new = StudentClass(class_name=op.payload["class_name"], students_count=op.payload["students_count"], school_id=school_id)
            db.add(new)
            db.flush()
            classes[new.class_name] = new
            result.created += 1
        return
    for op in sheet_plan.operations:
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(StudentClass, op.existing_id)
            if row is not None:
                row.students_count = op.payload["students_count"]
                classes[row.class_name] = row
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        new = StudentClass(class_name=op.payload["class_name"], students_count=op.payload["students_count"], school_id=school_id)
        db.add(new)
        db.flush()
        classes[new.class_name] = new
        result.created += 1


def _apply_teachers(db: Session, school_id: int, sheet_plan: SheetPlan, mode: ImportMode, teachers: dict[str, Teacher], result: CommitSheetResult) -> None:
    if mode == ImportMode.replace:
        for row in db.scalars(select(Teacher).where(Teacher.school_id == school_id)):
            db.delete(row)
        teachers.clear()
        db.flush()
        for op in sheet_plan.operations:
            new = Teacher(full_name=op.payload["full_name"], subjects=op.payload["subjects"], weekly_load_limit=op.payload["weekly_load_limit"], unavailable_days=op.payload["unavailable_days"], school_id=school_id)
            db.add(new)
            db.flush()
            teachers[new.full_name] = new
            result.created += 1
        return
    for op in sheet_plan.operations:
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(Teacher, op.existing_id)
            if row is not None:
                row.subjects = op.payload["subjects"]
                row.weekly_load_limit = op.payload["weekly_load_limit"]
                row.unavailable_days = op.payload["unavailable_days"]
                teachers[row.full_name] = row
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        new = Teacher(full_name=op.payload["full_name"], subjects=op.payload["subjects"], weekly_load_limit=op.payload["weekly_load_limit"], unavailable_days=op.payload["unavailable_days"], school_id=school_id)
        db.add(new)
        db.flush()
        teachers[new.full_name] = new
        result.created += 1


def _apply_classrooms(db: Session, school_id: int, sheet_plan: SheetPlan, mode: ImportMode, classrooms: dict[str, Classroom], result: CommitSheetResult) -> None:
    if mode == ImportMode.replace:
        for row in db.scalars(select(Classroom).where(Classroom.school_id == school_id)):
            db.delete(row)
        classrooms.clear()
        db.flush()
        for op in sheet_plan.operations:
            new = Classroom(room_number=op.payload["room_number"], capacity=op.payload["capacity"], specialization=op.payload["specialization"], school_id=school_id)
            db.add(new)
            db.flush()
            classrooms[new.room_number] = new
            result.created += 1
        return
    for op in sheet_plan.operations:
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(Classroom, op.existing_id)
            if row is not None:
                row.capacity = op.payload["capacity"]
                row.specialization = op.payload["specialization"]
                classrooms[row.room_number] = row
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        new = Classroom(room_number=op.payload["room_number"], capacity=op.payload["capacity"], specialization=op.payload["specialization"], school_id=school_id)
        db.add(new)
        db.flush()
        classrooms[new.room_number] = new
        result.created += 1


def _apply_group_flows(db: Session, school_id: int, sheet_plan: SheetPlan, mode: ImportMode, classes: dict[str, StudentClass], group_flows: dict[str, GroupFlow], result: CommitSheetResult) -> None:
    def _ids(names: list[str]) -> list[int]:
        ids: list[int] = []
        for member in names:
            cls = classes.get(member)
            if cls is not None:
                ids.append(cls.id)
        return ids

    if mode == ImportMode.replace:
        for row in db.scalars(select(GroupFlow).where(GroupFlow.school_id == school_id)):
            db.delete(row)
        group_flows.clear()
        db.flush()
        for op in sheet_plan.operations:
            new = GroupFlow(group_name=op.payload["group_name"], combined_classes=_ids(op.payload["combined_classes"]), school_id=school_id)
            db.add(new)
            db.flush()
            group_flows[new.group_name] = new
            result.created += 1
        return
    for op in sheet_plan.operations:
        ids = _ids(op.payload["combined_classes"])
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(GroupFlow, op.existing_id)
            if row is not None:
                row.combined_classes = ids
                group_flows[row.group_name] = row
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        new = GroupFlow(group_name=op.payload["group_name"], combined_classes=ids, school_id=school_id)
        db.add(new)
        db.flush()
        group_flows[new.group_name] = new
        result.created += 1


def _apply_curriculum(db: Session, school_id: int, sheet_plan: SheetPlan, mode: ImportMode, classes: dict[str, StudentClass], subjects: dict[str, Subject], result: CommitSheetResult) -> None:
    if mode == ImportMode.replace:
        for row in db.scalars(select(ClassSubjectHours).where(ClassSubjectHours.school_id == school_id)):
            db.delete(row)
        db.flush()
        for op in sheet_plan.operations:
            cls = classes.get(op.payload["class_name"])
            subj = subjects.get(op.payload["subject_name"])
            if cls is None or subj is None:
                result.skipped += 1
                continue
            db.add(ClassSubjectHours(school_id=school_id, class_id=cls.id, subject_id=subj.id, hours_per_week=op.payload["hours_per_week"]))
            result.created += 1
        return
    for op in sheet_plan.operations:
        cls = classes.get(op.payload["class_name"])
        subj = subjects.get(op.payload["subject_name"])
        if cls is None or subj is None:
            result.skipped += 1
            continue
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(ClassSubjectHours, op.existing_id)
            if row is not None:
                row.hours_per_week = op.payload["hours_per_week"]
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        db.add(ClassSubjectHours(school_id=school_id, class_id=cls.id, subject_id=subj.id, hours_per_week=op.payload["hours_per_week"]))
        result.created += 1


def _apply_schedule(
    db: Session,
    school_id: int,
    sheet_plan: SheetPlan,
    mode: ImportMode,
    classes: dict[str, StudentClass],
    subjects: dict[str, Subject],
    teachers: dict[str, Teacher],
    classrooms: dict[str, Classroom],
    lesson_slots: dict[tuple[int, int], LessonSlot],
    group_flows: dict[str, GroupFlow],
    result: CommitSheetResult,
) -> None:
    def _resolve(op: _RowOp) -> dict[str, Any] | None:
        cls = classes.get(op.payload["class_name"])
        subj = subjects.get(op.payload["subject_name"])
        tch = teachers.get(op.payload["teacher_full_name"])
        room = classrooms.get(op.payload["room_number"])
        slot = lesson_slots.get((op.payload["day_of_week"], op.payload["lesson_number"]))
        if not (cls and subj and tch and room and slot):
            return None
        group_id: int | None = None
        if op.payload["is_grouped"] and op.payload["group_name"]:
            flow = group_flows.get(op.payload["group_name"])
            if flow is None:
                return None
            group_id = flow.id
        return {
            "class_id": cls.id,
            "subject_id": subj.id,
            "teacher_id": tch.id,
            "classroom_id": room.id,
            "lesson_slot_id": slot.id,
            "is_grouped": bool(op.payload["is_grouped"]),
            "group_id": group_id,
            "school_id": school_id,
        }

    if mode == ImportMode.replace:
        for row in db.scalars(select(ScheduleItem).where(ScheduleItem.school_id == school_id)):
            db.delete(row)
        db.flush()
        for op in sheet_plan.operations:
            payload = _resolve(op)
            if payload is None:
                result.skipped += 1
                continue
            db.add(ScheduleItem(**payload))
            result.created += 1
        return
    for op in sheet_plan.operations:
        payload = _resolve(op)
        if payload is None:
            result.skipped += 1
            continue
        if op.action == "update" and mode != ImportMode.append and op.existing_id:
            row = db.get(ScheduleItem, op.existing_id)
            if row is not None:
                row.subject_id = payload["subject_id"]
                row.teacher_id = payload["teacher_id"]
                row.classroom_id = payload["classroom_id"]
                row.is_grouped = payload["is_grouped"]
                row.group_id = payload["group_id"]
                result.updated += 1
                continue
        if op.action == "update" and mode == ImportMode.append:
            result.skipped += 1
            continue
        db.add(ScheduleItem(**payload))
        result.created += 1


# ---------------------------------------------------------------------------
# Helper aggregator for downstream API layer
# ---------------------------------------------------------------------------


def issue_counts(issues: list[ImportIssue]) -> tuple[int, int]:
    error = sum(1 for issue in issues if issue.severity == IssueSeverity.error.value)
    warning = sum(1 for issue in issues if issue.severity == IssueSeverity.warning.value)
    return error, warning


__all__ = [
    "apply_plan",
    "build_plan",
    "build_template_workbook",
    "ImportPlan",
    "issue_counts",
    "load_workbook_from_bytes",
    "plan_to_summary",
    "SheetPlan",
]
