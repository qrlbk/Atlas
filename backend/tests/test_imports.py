"""Tests for /imports/template, /imports/validate and /imports/commit endpoints."""

from __future__ import annotations

import io
from datetime import time

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.security import get_password_hash
from app.models.entities import (
    ClassSubjectHours,
    Classroom,
    ClassroomSpecialization,
    LessonSlot,
    ScheduleItem,
    School,
    StudentClass,
    Subject,
    Teacher,
    User,
    UserRole,
)


def _seed_school(db: Session) -> dict:
    school = School(name="Main School", address="Main St")
    db.add(school)
    db.flush()

    manager = User(
        email="mgr@example.com",
        full_name="Manager",
        password_hash=get_password_hash("password"),
        role=UserRole.school_manager,
        school_id=school.id,
    )
    other_school = School(name="Other", address="Other St")
    db.add(other_school)
    db.flush()
    other_manager = User(
        email="other@example.com",
        full_name="Other Manager",
        password_hash=get_password_hash("password"),
        role=UserRole.school_manager,
        school_id=other_school.id,
    )
    db.add_all([manager, other_manager])
    db.commit()
    return {"school": school, "other_school": other_school}


def _token(api_client: TestClient, email: str) -> dict:
    r = api_client.post("/auth/login", json={"email": email, "password": "password"})
    assert r.status_code == 200, r.text
    return {"Authorization": f"Bearer {r.json()['access_token']}"}


def _build_workbook(rows: dict[str, list[list[object]]]) -> bytes:
    """Build an xlsx workbook from {sheet_name: [headers, *rows]}."""

    wb = Workbook()
    default = wb.active
    wb.remove(default)
    for sheet_name, sheet_rows in rows.items():
        ws = wb.create_sheet(title=sheet_name)
        for row in sheet_rows:
            ws.append(row)
    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def test_template_download_returns_xlsx(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    headers = _token(api_client, "mgr@example.com")
    r = api_client.get(f"/imports/template?school_id={ctx['school'].id}", headers=headers)
    assert r.status_code == 200
    assert r.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    wb = load_workbook(io.BytesIO(r.content))
    assert "Teachers" in wb.sheetnames
    assert "Schedule" in wb.sheetnames


def test_template_requires_school_scope(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    headers = _token(api_client, "other@example.com")
    r = api_client.get(f"/imports/template?school_id={ctx['school'].id}", headers=headers)
    assert r.status_code == 403


def test_validate_basic_upsert(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    school_id = ctx["school"].id
    headers = _token(api_client, "mgr@example.com")

    payload = _build_workbook({
        "Subjects": [
            ["name", "requires_special_room", "required_specialization"],
            ["Math", False, ""],
            ["Chemistry", True, "chemistry_lab"],
        ],
        "LessonSlots": [
            ["day_of_week", "lesson_number", "start_time", "end_time"],
            [1, 1, "08:00", "08:45"],
            [1, 2, "08:55", "09:40"],
        ],
        "Classes": [
            ["class_name", "students_count"],
            ["9A", 24],
        ],
        "Teachers": [
            ["full_name", "subjects", "weekly_load_limit", "unavailable_days"],
            ["Ada Lovelace", "Math", 20, "6,7"],
        ],
        "Classrooms": [
            ["room_number", "capacity", "specialization"],
            ["101", 30, "standard"],
        ],
        "Curriculum": [
            ["class_name", "subject_name", "hours_per_week"],
            ["9A", "Math", 4],
        ],
        "Schedule": [
            ["class_name", "subject_name", "teacher_full_name", "room_number", "day_of_week", "lesson_number", "is_grouped", "group_name"],
            ["9A", "Math", "Ada Lovelace", "101", 1, 1, False, ""],
        ],
    })

    r = api_client.post(
        "/imports/validate",
        data={"school_id": school_id},
        files={"file": ("data.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["can_commit"] is True
    assert body["summary"]["error_count"] == 0
    summary_by_sheet = {row["sheet"]: row for row in body["summary"]["sheets"]}
    assert summary_by_sheet["Subjects"]["rows_to_create"] == 2
    assert summary_by_sheet["Schedule"]["rows_to_replace"] == 1  # default mode for Schedule is replace


def test_validate_reports_unknown_class_in_schedule(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    school_id = ctx["school"].id
    headers = _token(api_client, "mgr@example.com")

    payload = _build_workbook({
        "Subjects": [
            ["name", "requires_special_room", "required_specialization"],
            ["Math", False, ""],
        ],
        "LessonSlots": [
            ["day_of_week", "lesson_number", "start_time", "end_time"],
            [1, 1, "08:00", "08:45"],
        ],
        "Teachers": [
            ["full_name", "subjects", "weekly_load_limit", "unavailable_days"],
            ["Ada", "Math", 0, ""],
        ],
        "Classrooms": [
            ["room_number", "capacity", "specialization"],
            ["101", 30, "standard"],
        ],
        "Schedule": [
            ["class_name", "subject_name", "teacher_full_name", "room_number", "day_of_week", "lesson_number", "is_grouped", "group_name"],
            ["NOPE", "Math", "Ada", "101", 1, 1, False, ""],
        ],
    })

    r = api_client.post(
        "/imports/validate",
        data={"school_id": school_id},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["can_commit"] is False
    codes = {issue["code"] for issue in body["issues"]}
    assert "unknown_class" in codes


def test_commit_creates_entities(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    school_id = ctx["school"].id
    headers = _token(api_client, "mgr@example.com")

    payload = _build_workbook({
        "Subjects": [
            ["name", "requires_special_room", "required_specialization"],
            ["Math", False, ""],
        ],
        "LessonSlots": [
            ["day_of_week", "lesson_number", "start_time", "end_time"],
            [1, 1, "08:00", "08:45"],
        ],
        "Classes": [
            ["class_name", "students_count"],
            ["9A", 24],
        ],
        "Teachers": [
            ["full_name", "subjects", "weekly_load_limit", "unavailable_days"],
            ["Ada Lovelace", "Math", 20, ""],
        ],
        "Classrooms": [
            ["room_number", "capacity", "specialization"],
            ["101", 30, "standard"],
        ],
        "Curriculum": [
            ["class_name", "subject_name", "hours_per_week"],
            ["9A", "Math", 4],
        ],
        "Schedule": [
            ["class_name", "subject_name", "teacher_full_name", "room_number", "day_of_week", "lesson_number", "is_grouped", "group_name"],
            ["9A", "Math", "Ada Lovelace", "101", 1, 1, False, ""],
        ],
    })

    r = api_client.post(
        "/imports/commit",
        data={"school_id": school_id},
        files={"file": ("data.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["committed"] is True

    db_session.expire_all()
    assert db_session.query(Subject).filter_by(name="Math").one() is not None
    cls = db_session.query(StudentClass).filter_by(school_id=school_id, class_name="9A").one()
    assert cls.students_count == 24
    teacher = db_session.query(Teacher).filter_by(school_id=school_id, full_name="Ada Lovelace").one()
    assert teacher.subjects == ["Math"]
    items = list(db_session.query(ScheduleItem).filter_by(school_id=school_id))
    assert len(items) == 1


def test_commit_replace_schedule_only_for_this_school(api_client: TestClient, db_session: Session):
    """Schedule replace must not delete schedule items of other schools."""

    ctx = _seed_school(db_session)
    school_id = ctx["school"].id
    other_id = ctx["other_school"].id
    headers = _token(api_client, "mgr@example.com")

    subj = Subject(name="Math", requires_special_room=False, required_specialization=None)
    slot = LessonSlot(day_of_week=1, lesson_number=1, start_time=time(8, 0), end_time=time(8, 45))
    cls_main = StudentClass(class_name="9A", students_count=20, school_id=school_id)
    cls_other = StudentClass(class_name="9X", students_count=20, school_id=other_id)
    tch_main = Teacher(full_name="T", subjects=["Math"], weekly_load_limit=0, unavailable_days=[], school_id=school_id)
    tch_other = Teacher(full_name="T2", subjects=["Math"], weekly_load_limit=0, unavailable_days=[], school_id=other_id)
    room_main = Classroom(room_number="101", capacity=30, specialization=ClassroomSpecialization.standard, school_id=school_id)
    room_other = Classroom(room_number="201", capacity=30, specialization=ClassroomSpecialization.standard, school_id=other_id)
    db_session.add_all([subj, slot, cls_main, cls_other, tch_main, tch_other, room_main, room_other])
    db_session.flush()

    db_session.add_all([
        ScheduleItem(class_id=cls_main.id, subject_id=subj.id, teacher_id=tch_main.id, classroom_id=room_main.id, lesson_slot_id=slot.id, is_grouped=False, group_id=None, school_id=school_id),
        ScheduleItem(class_id=cls_other.id, subject_id=subj.id, teacher_id=tch_other.id, classroom_id=room_other.id, lesson_slot_id=slot.id, is_grouped=False, group_id=None, school_id=other_id),
    ])
    db_session.commit()

    payload = _build_workbook({
        "Schedule": [
            ["class_name", "subject_name", "teacher_full_name", "room_number", "day_of_week", "lesson_number", "is_grouped", "group_name"],
        ],
    })
    r = api_client.post(
        "/imports/commit",
        data={"school_id": school_id, "modes": '{"Schedule": "replace"}'},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 200, r.text

    db_session.expire_all()
    remaining = list(db_session.query(ScheduleItem))
    assert len(remaining) == 1
    assert remaining[0].school_id == other_id


def test_commit_rejects_workbook_with_errors(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    school_id = ctx["school"].id
    headers = _token(api_client, "mgr@example.com")

    payload = _build_workbook({
        "Subjects": [
            ["name", "requires_special_room", "required_specialization"],
            ["Math", False, ""],
        ],
        "Schedule": [
            ["class_name", "subject_name", "teacher_full_name", "room_number", "day_of_week", "lesson_number", "is_grouped", "group_name"],
            ["NOPE", "Math", "Nobody", "101", 1, 1, False, ""],
        ],
    })

    r = api_client.post(
        "/imports/commit",
        data={"school_id": school_id},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["committed"] is False
    assert body["applied"] == []
    db_session.expire_all()
    assert db_session.query(ScheduleItem).count() == 0
    assert db_session.query(Subject).count() == 0


def test_validate_blocks_cross_school_scope(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    headers = _token(api_client, "other@example.com")
    payload = _build_workbook({"Subjects": [["name"], ["Math"]]})
    r = api_client.post(
        "/imports/validate",
        data={"school_id": ctx["school"].id},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 403


def test_validate_viewer_forbidden(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    viewer = User(
        email="viewer@example.com",
        full_name="V",
        password_hash=get_password_hash("password"),
        role=UserRole.viewer,
        school_id=ctx["school"].id,
    )
    db_session.add(viewer)
    db_session.commit()
    headers = _token(api_client, "viewer@example.com")
    payload = _build_workbook({"Subjects": [["name"], ["Math"]]})
    r = api_client.post(
        "/imports/validate",
        data={"school_id": ctx["school"].id},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 403


def test_append_mode_skips_existing_rows(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    school_id = ctx["school"].id
    headers = _token(api_client, "mgr@example.com")

    db_session.add(Teacher(full_name="Ada", subjects=["Math"], weekly_load_limit=10, unavailable_days=[], school_id=school_id))
    db_session.commit()

    payload = _build_workbook({
        "Teachers": [
            ["full_name", "subjects", "weekly_load_limit", "unavailable_days"],
            ["Ada", "Physics", 20, ""],
            ["Bob", "Math", 15, ""],
        ],
    })
    r = api_client.post(
        "/imports/commit",
        data={"school_id": school_id, "modes": '{"Teachers": "append"}'},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 200
    body = r.json()
    assert body["committed"] is True
    db_session.expire_all()
    ada = db_session.query(Teacher).filter_by(school_id=school_id, full_name="Ada").one()
    assert ada.subjects == ["Math"]  # untouched
    assert ada.weekly_load_limit == 10
    bob = db_session.query(Teacher).filter_by(school_id=school_id, full_name="Bob").one()
    assert bob.weekly_load_limit == 15


def test_invalid_modes_payload(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    headers = _token(api_client, "mgr@example.com")
    payload = _build_workbook({"Subjects": [["name"], ["Math"]]})
    r = api_client.post(
        "/imports/validate",
        data={"school_id": ctx["school"].id, "modes": '{"Subjects": "explode"}'},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 400
    assert r.json().get("code") == "errors.importInvalidModes"


def test_invalid_xlsx_returns_400(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    headers = _token(api_client, "mgr@example.com")
    r = api_client.post(
        "/imports/validate",
        data={"school_id": ctx["school"].id},
        files={"file": ("d.xlsx", b"not really a workbook", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 400
    assert r.json().get("code") == "errors.importInvalidWorkbook"


def test_curriculum_update_existing_row(api_client: TestClient, db_session: Session):
    ctx = _seed_school(db_session)
    school_id = ctx["school"].id
    headers = _token(api_client, "mgr@example.com")

    subj = Subject(name="Math", requires_special_room=False, required_specialization=None)
    cls = StudentClass(class_name="9A", students_count=24, school_id=school_id)
    db_session.add_all([subj, cls])
    db_session.flush()
    db_session.add(ClassSubjectHours(school_id=school_id, class_id=cls.id, subject_id=subj.id, hours_per_week=2))
    db_session.commit()

    payload = _build_workbook({
        "Curriculum": [
            ["class_name", "subject_name", "hours_per_week"],
            ["9A", "Math", 5],
        ],
    })
    r = api_client.post(
        "/imports/commit",
        data={"school_id": school_id},
        files={"file": ("d.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        headers=headers,
    )
    assert r.status_code == 200, r.text
    body = r.json()
    assert body["committed"] is True
    db_session.expire_all()
    row = db_session.query(ClassSubjectHours).filter_by(school_id=school_id).one()
    assert row.hours_per_week == 5
